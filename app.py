from flask import Flask, render_template, request, send_from_directory
from openpyxl import Workbook, load_workbook
import os
import dropbox

app = Flask(__name__)

# Access Token di Dropbox (inserisci il tuo token qui)
DROPBOX_ACCESS_TOKEN = 'sl.CBwlaIugRUpVnh_mTAIqLNuii95RmSAwECy623bPfCUtuDyEDfNwe-QOJldhY63546YDFkeKsl-qUAFRFDkfhnj3cmmD7mG3Ij2iIWjiQhwKxuXAhdRleK4p0yXRlXBsSMOwbJye7AFW'

# Test dell'accesso a Dropbox
def test_dropbox_access():
    try:
        # Creazione di una connessione con Dropbox
        dbx = dropbox.Dropbox(DROPBOX_ACCESS_TOKEN)
        # Prova a ottenere informazioni sull'account
        account_info = dbx.users_get_current_account()
        print(f"Accesso a Dropbox riuscito! Account: {account_info.name.display_name}")
    except dropbox.exceptions.AuthError as e:
        print(f"Errore di autenticazione: {e}")

# Esegui il test di accesso a Dropbox
test_dropbox_access()

# Percorso del file Excel
EXCEL_FILE = "7tmp7adesioni.xlsx"

# Creazione iniziale del file Excel se non esiste
if not os.path.exists(EXCEL_FILE):
    print("Creazione del file Excel...")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Adesioni"
    sheet.append(["Cognome", "Nome", "Data di Nascita", "Luogo di Nascita", "Codice Fiscale", "Email", "Cellulare"])
    workbook.save(EXCEL_FILE)
    print("File Excel creato!")
else:
    print("File Excel già esistente.")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    # Raccogliere i dati inviati dal modulo
    cognome = request.form.get('surname')
    nome = request.form.get('name')
    data_nascita = request.form.get('birthdate')
    luogo_nascita = request.form.get('birthplace')
    codice_fiscale = request.form.get('fiscalcode')
    email = request.form.get('email')
    cellulare = request.form.get('phone')

    print("Dati ricevuti dal modulo:", cognome, nome, data_nascita, luogo_nascita, codice_fiscale, email, cellulare)

    try:
        # Aprire o creare il file Excel
        if not os.path.exists(EXCEL_FILE):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Adesioni"
            sheet.append(["Cognome", "Nome", "Data di Nascita", "Luogo di Nascita", "Codice Fiscale", "Email", "Cellulare"])
        else:
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active

        # Aggiungere i dati
        sheet.append([cognome, nome, data_nascita, luogo_nascita, codice_fiscale, email, cellulare])
        workbook.save(EXCEL_FILE)
        print("Dati salvati con successo!")

        # Caricare il file su Dropbox
        upload_to_dropbox(EXCEL_FILE)
    except Exception as e:
        print(f"Errore durante il salvataggio: {e}")
        return f"Errore durante il salvataggio: {e}"

    return "Iscrizione completata con successo!"

# Funzione per caricare il file su Dropbox
def upload_to_dropbox(file_path):
    # Creazione di una connessione con Dropbox
    dbx = dropbox.Dropbox(DROPBOX_ACCESS_TOKEN)

    try:
        # Aprire il file e caricarlo su Dropbox
        with open(file_path, "rb") as file:
            dropbox_path = f"/cartella_destinazione/{os.path.basename(file_path)}"  # Modifica qui se necessario
            dbx.files_upload(file.read(), dropbox_path, mute=True)
            print(f"File {file_path} caricato su Dropbox!")
    except Exception as e:
        print(f"Errore durante il caricamento su Dropbox: {e}")

@app.route('/download', methods=['GET'])
def download_file():
    # Invia il file Excel dal server al client
    return send_from_directory(directory=os.getcwd(), filename=EXCEL_FILE)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)), debug=True)
